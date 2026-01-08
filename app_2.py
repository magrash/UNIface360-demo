from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, Response, send_file
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3
from os.path import basename
import os
from datetime import datetime, timedelta
import csv
from io import StringIO, BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
from PIL import Image as PILImage
import cv2
import threading
import atexit
from flask_mail import Mail, Message
import logging
from calendar import monthrange
import random
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.backends.backend_agg import FigureCanvasAgg
import urllib.request
import requests

# Initialize Flask app

# Function to generate employee data with real names
def generate_employee_data():
    """Generate 500 employees with real English names and photos"""
    first_names_english = [
        "James", "John", "Robert", "Michael", "William", "David", "Richard", "Joseph", "Thomas", "Charles",
        "Christopher", "Daniel", "Matthew", "Anthony", "Mark", "Donald", "Steven", "Paul", "Andrew", "Joshua",
        "Kenneth", "Kevin", "Brian", "George", "Timothy", "Ronald", "Jason", "Edward", "Jeffrey", "Ryan",
        "Jacob", "Gary", "Nicholas", "Eric", "Jonathan", "Stephen", "Larry", "Justin", "Scott", "Brandon",
        "Benjamin", "Samuel", "Frank", "Gregory", "Raymond", "Alexander", "Patrick", "Jack", "Dennis", "Jerry",
        "Peter", "Henry", "Carl", "Arthur", "Alan", "Lawrence", "Roger", "Keith", "Jeremy", "Terry",
        "Sean", "Gerald", "Christian", "Jose", "Adam", "Nathan", "Zachary", "Kyle", "Noah", "Ethan",
        "Mason", "Logan", "Lucas", "Aiden", "Jackson", "Mason", "Carter", "Owen", "Wyatt", "Grayson",
        "Leo", "Julian", "Luke", "Hunter", "Connor", "Eli", "Aaron", "Caleb", "Isaac", "Landon",
        "Adrian", "Evan", "Nolan", "Tyler", "Colin", "Jaxon", "Brayden", "Dominic", "Austin", "Jordan"
    ]
    
    last_names_english = [
        "Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller", "Davis", "Rodriguez", "Martinez",
        "Hernandez", "Lopez", "Wilson", "Anderson", "Thomas", "Taylor", "Moore", "Jackson", "Martin", "Lee",
        "Thompson", "White", "Harris", "Sanchez", "Clark", "Ramirez", "Lewis", "Robinson", "Walker", "Young",
        "Allen", "King", "Wright", "Scott", "Torres", "Nguyen", "Hill", "Flores", "Green", "Adams",
        "Nelson", "Baker", "Hall", "Rivera", "Campbell", "Mitchell", "Carter", "Roberts", "Gomez", "Phillips",
        "Evans", "Turner", "Diaz", "Parker", "Cruz", "Edwards", "Collins", "Reyes", "Stewart", "Morris",
        "Rogers", "Reed", "Cook", "Morgan", "Bell", "Murphy", "Bailey", "Rivera", "Cooper", "Richardson",
        "Cox", "Howard", "Ward", "Torres", "Peterson", "Gray", "Ramirez", "James", "Watson", "Brooks",
        "Kelly", "Sanders", "Price", "Bennett", "Wood", "Barnes", "Ross", "Henderson", "Coleman", "Jenkins",
        "Perry", "Powell", "Long", "Patterson", "Hughes", "Flores", "Washington", "Butler", "Simmons", "Foster"
    ]
    
    departments = ["Drilling", "Operations", "Production", "HSE", "Maintenance", "Logistics"]
    positions = ["HSE Officer", "Engineer", "Technician", "Operator", "Supervisor", "Analyst", "Manager", "Coordinator"]
    supervisors = ["S. Fathy", "H. Mostafa", "M. Farouk", "I. Khan", "L. Kamal", "R. Nabil", "A. Hassan", "K. Ali"]
    
    employees = []
    for i in range(1, 501):
        # Use only English names
        first_name = random.choice(first_names_english)
        last_name = random.choice(last_names_english)
        full_name = f"{first_name} {last_name}"
        
        # Generate hire date between 2020 and 2024
        year = random.randint(2020, 2024)
        month = random.randint(1, 12)
        day = random.randint(1, 28)
        hire_date = f"{year}-{month:02d}-{day:02d}"
        
        # Photo URL using pravatar with unique ID
        photo_url = f"https://i.pravatar.cc/150?img={i % 70 + 1}"  # Cycle through 70 different avatars
        
        employees.append({
            "id": f"E{i:03d}",
            "name": full_name,
            "department": random.choice(departments),
            "position": random.choice(positions),
            "supervisor": random.choice(supervisors),
            "hire_date": hire_date,
            "photo_url": photo_url
        })
    
    return employees

# Cache employee data
_employee_data_cache = None

def get_employee_data():
    """Get cached employee data or generate if not exists"""
    global _employee_data_cache
    if _employee_data_cache is None:
        _employee_data_cache = generate_employee_data()
    return _employee_data_cache

def reset_employee_data_cache():
    """Reset the employee data cache to force regeneration"""
    global _employee_data_cache
    _employee_data_cache = None

def get_employee_by_id(employee_id):
    """Get employee by ID"""
    employees = get_employee_data()
    for emp in employees:
        if emp["id"] == employee_id:
            return emp
    return None

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "your-secret-key-here")
app.jinja_env.filters['basename'] = basename

# Configure Flask-Mail
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'mustafamagdy2002@gmail.com'  # Replace with your email
app.config['MAIL_PASSWORD'] = 'fczs xzgx xfmg ecby'  # Replace with the generated app password

mail = Mail(app)  # Initialize Mail after configuration

# Ensure mail.logger is not None before setting the logging level
if mail.logger:
    mail.logger.setLevel(logging.DEBUG)

# Base directory for the project (adjust if needed)
BASE_DIR = r"C:\SW\Camera attendence"

# Setup Flask-Login

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

# User database
def init_user_db():
    conn = sqlite3.connect("users.db")
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            password TEXT
        )
    """)
    cursor.execute("SELECT * FROM users WHERE username = ?", ("admin",))
    if not cursor.fetchone():
        cursor.execute("INSERT INTO users (username, password) VALUES (?, ?)",
                      ("admin", generate_password_hash("admin123")))
    # Demo employee account for frontend testing
    cursor.execute("SELECT * FROM users WHERE username = ?", ("employee",))
    if not cursor.fetchone():
        cursor.execute("INSERT INTO users (username, password) VALUES (?, ?)",
                      ("employee", generate_password_hash("emp123")))
    # HSE roles
    cursor.execute("SELECT * FROM users WHERE username = ?", ("admin_hse",))
    if not cursor.fetchone():
        cursor.execute("INSERT INTO users (username, password) VALUES (?, ?)",
                      ("admin_hse", generate_password_hash("admin123")))
    cursor.execute("SELECT * FROM users WHERE username = ?", ("user_hse",))
    if not cursor.fetchone():
        cursor.execute("INSERT INTO users (username, password) VALUES (?, ?)",
                      ("user_hse", generate_password_hash("emp123")))
    conn.commit()
    conn.close()

# Initialize tracking database with sample data
def init_tracking_db():
    conn = sqlite3.connect("tracking.db")
    cursor = conn.cursor()
    
    # Check if 'floor' column exists and rename it to 'location'
    cursor.execute("PRAGMA table_info(logs)")
    columns = [col[1] for col in cursor.fetchall()]
    if 'floor' in columns and 'location' not in columns:
        cursor.execute("ALTER TABLE logs RENAME COLUMN floor TO location")
    
    # Create logs table if it doesn't exist
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS logs (
            name TEXT,
            location TEXT,
            time TEXT,
            image_path TEXT
        )
    """)
    
    # Create excuses table if it doesn't exist
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS excuses (
            name TEXT,
            date TEXT,
            hours INTEGER,
            reason TEXT,
            approved_by TEXT,
            created_at TEXT
        )
    """)
    
    # Update existing data to use new room labels
    cursor.execute("UPDATE logs SET location = 'Meeting Room' WHERE location = 'A'")
    cursor.execute("UPDATE logs SET location = 'Main Room' WHERE location = 'B' OR location = 'C' OR location LIKE 'Floor%'")
    cursor.execute("UPDATE logs SET location = 'Room 1' WHERE location = 'R1'")
    cursor.execute("UPDATE logs SET location = 'Room 2' WHERE location = 'R2'")
    cursor.execute("UPDATE logs SET location = 'Room 3' WHERE location = 'R3'")
    cursor.execute("UPDATE logs SET location = 'Room 4' WHERE location = 'R4'")
    cursor.execute("UPDATE logs SET location = 'Kitchen' WHERE location = 'K'")
    
    conn.commit()
    conn.close()

init_user_db()
init_tracking_db()

class User(UserMixin):
    def __init__(self, id, username):
        self.id = id
        self.username = username

@login_manager.user_loader
def load_user(user_id):
    conn = sqlite3.connect("users.db")
    cursor = conn.execute("SELECT id, username FROM users WHERE id = ?", (user_id,))
    user = cursor.fetchone()
    conn.close()
    if user:
        return User(user[0], user[1])
    return None

@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        # Redirect to correct dashboard by username
        u = current_user.username
        if u == "admin":
            return redirect(url_for("hr_home"))
        if u == "employee":
            return redirect(url_for("employee_dashboard"))
        if u == "admin_hse":
            return redirect(url_for("admin_hse_dashboard"))
        if u == "user_hse":
            return redirect(url_for("employee_hse_dashboard"))
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        conn = sqlite3.connect("users.db")
        cursor = conn.execute("SELECT id, username, password FROM users WHERE username = ?", (username,))
        user = cursor.fetchone()
        conn.close()
        if user and check_password_hash(user[2], password):
            login_user(User(user[0], user[1]))
            if username == "admin":
                return redirect(url_for("hr_home"))
            if username == "employee":
                return redirect(url_for("employee_dashboard"))
            if username == "admin_hse":
                return redirect(url_for("admin_hse_dashboard"))
            if username == "user_hse":
                return redirect(url_for("employee_hse_dashboard"))
            return redirect(url_for("dashboard"))
        flash("Invalid username or password", "error")
    return render_template("login_2.html")

@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("You have been logged out.", "success")
    return redirect(url_for("login"))

@app.route("/home")
@app.route("/")
def home():
    return render_template("home.html")

@app.route("/support")
def support():
    return render_template("support.html")

@app.route("/send_support_email", methods=["POST"])
def send_support_email():
    name = request.form.get("name")
    email = request.form.get("email")
    message = request.form.get("message")

    # Compose the email
    msg = Message(
        subject="Support Request from Uniface360",
        sender=app.config['MAIL_USERNAME'],
        recipients=["mostafa.magdy@petrochoice.org"],  # Default recipient
        body=f"Name: {name}\nEmail: Admin@Uniface360\n\nMessage:\n{message}"
    )

    # Send the email
    try:
        mail.send(msg)
        flash("Your message has been sent successfully!", "success")
    except Exception as e:
        flash("Failed to send your message. Please try again later.", "error")
        print(f"Error: {e}")
    
    return redirect(url_for("support"))


@app.route("/request_demo", methods=["GET"])
def request_demo():
    return render_template("request_demo.html")


@app.route("/send_demo_request", methods=["POST"])
def send_demo_request():
    name = request.form.get("name")
    company = request.form.get("company")
    email = request.form.get("email", "Not provided")
    phone = request.form.get("phone", "Not provided")
    building_size = request.form.get("building_size", "Not provided")
    message = request.form.get("message", "")

    # Compose the email
    msg = Message(
        subject="Demo Request for Uniface360",
        sender=app.config['MAIL_USERNAME'],
        recipients=["mostafa.magdy@petrochoice.org"],  # Default recipient
        body=f"Demo Request Details:\n\nName: {name}\nCompany: {company}\nEmail: {email}\nPhone: {phone}\nBuilding Size: {building_size}\n\nAdditional Information:\n{message}"
    )

    # Send the email
    try:
        mail.send(msg)
        flash("Your demo request has been submitted successfully! Our team will contact you shortly.", "success")
    except Exception as e:
        flash("Failed to submit your demo request. Please try again later.", "danger")
        print(f"Error: {e}")
    
    return redirect(url_for("request_demo"))

@app.route("/dashboard")
@login_required
def dashboard():
    conn = sqlite3.connect("tracking.db")
    # Location data: Latest known location for each person
    cursor = conn.execute("""
        SELECT l.location, COUNT(DISTINCT l.name) as count, GROUP_CONCAT(l.name) as names
        FROM logs l
        INNER JOIN (
            SELECT name, MAX(time) as max_time
            FROM logs
            WHERE name != 'Unknown'
            GROUP BY name
        ) latest ON l.name = latest.name AND l.time = latest.max_time
        GROUP BY l.location
    """)
    location_data = {row[0]: {"count": row[1], "names": row[2].split(",") if row[2] else []} for row in cursor}
    
    # Get the latest location for each person (for the map)
    cursor = conn.execute("""
        SELECT name, location, MAX(time) as last_seen
        FROM logs
        WHERE name != 'Unknown'
        GROUP BY name
    """)
    people = [{"name": row[0], "location": row[1], "last_seen": row[2]} for row in cursor]
    
    # Attendance analysis: Analyze the most recent day for each person
    cursor = conn.execute("""
        SELECT name, date, MIN(time) as first_seen, MAX(time) as last_seen
        FROM (
            SELECT name, time, SUBSTR(time, 1, 10) as date
            FROM logs
            WHERE name != 'Unknown'
        ) sub
        GROUP BY name, date
        HAVING date = (
            SELECT SUBSTR(MAX(time), 1, 10)
            FROM logs l2
            WHERE l2.name = sub.name
        )
    """)
    attendance = []
    for row in cursor:
        name, date, first, last = row
        first_time = datetime.strptime(first, "%Y-%m-%d %H:%M:%S")
        last_time = datetime.strptime(last, "%Y-%m-%d %H:%M:%S")
        late = first_time.hour > 8 or (first_time.hour == 8 and first_time.minute > 0)
        early_leave = last_time.hour < 16
        attendance.append({
            "name": name,
            "date": date,
            "first_seen": first,
            "last_seen": last,
            "late": late,
            "early_leave": early_leave
        })
    
    # Total detections per person
    cursor = conn.execute("SELECT name, COUNT(*) as count FROM logs WHERE name != 'Unknown' GROUP BY name")
    person_counts = {row[0]: row[1] for row in cursor}
    # Detections per location
    cursor = conn.execute("SELECT location, COUNT(*) as count FROM logs WHERE name != 'Unknown' GROUP BY location")
    location_counts = {row[0]: row[1] for row in cursor}
    # Recent activity
    cursor = conn.execute("SELECT name, location, time FROM logs WHERE name != 'Unknown' ORDER BY time DESC LIMIT 5")
    recent_logs = [{"name": row[0], "location": row[1], "time": row[2]} for row in cursor]
    conn.close()
    return render_template("dashboard_2.html", 
                         floor_data=location_data, 
                         attendance=attendance,
                         person_counts=person_counts, 
                         floor_counts=location_counts, 
                         recent_logs=recent_logs,
                         people=people)

@app.route("/admin")
@login_required
def admin_dashboard():
    # Normal Admin dashboard (non-HSE)
    if current_user.username != "admin":
        return redirect(url_for("employee_dashboard"))
    
    conn = sqlite3.connect("tracking.db")
    # Location data: Latest known location for each person
    cursor = conn.execute("""
        SELECT l.location, COUNT(DISTINCT l.name) as count, GROUP_CONCAT(l.name) as names
        FROM logs l
        INNER JOIN (
            SELECT name, MAX(time) as max_time
            FROM logs
            WHERE name != 'Unknown'
            GROUP BY name
        ) latest ON l.name = latest.name AND l.time = latest.max_time
        GROUP BY l.location
    """)
    location_data = {row[0]: {"count": row[1], "names": row[2].split(",") if row[2] else []} for row in cursor}
    
    # Get the latest location for each person (for the map)
    cursor = conn.execute("""
        SELECT name, location, MAX(time) as last_seen
        FROM logs
        WHERE name != 'Unknown'
        GROUP BY name
    """)
    people = [{"name": row[0], "location": row[1], "last_seen": row[2]} for row in cursor]
    
    # Attendance analysis: Analyze the most recent day for each person
    cursor = conn.execute("""
        SELECT name, date, MIN(time) as first_seen, MAX(time) as last_seen
        FROM (
            SELECT name, time, SUBSTR(time, 1, 10) as date
            FROM logs
            WHERE name != 'Unknown'
        ) sub
        GROUP BY name, date
        HAVING date = (
            SELECT SUBSTR(MAX(time), 1, 10)
            FROM logs l2
            WHERE l2.name = sub.name
        )
    """)
    attendance = []
    for row in cursor:
        name, date, first, last = row
        first_time = datetime.strptime(first, "%Y-%m-%d %H:%M:%S")
        last_time = datetime.strptime(last, "%Y-%m-%d %H:%M:%S")
        late = first_time.hour > 8 or (first_time.hour == 8 and first_time.minute > 0)
        early_leave = last_time.hour < 16
        attendance.append({
            "name": name,
            "date": date,
            "first_seen": first,
            "last_seen": last,
            "late": late,
            "early_leave": early_leave
        })
    
    # Total detections per person
    cursor = conn.execute("SELECT name, COUNT(*) as count FROM logs WHERE name != 'Unknown' GROUP BY name")
    person_counts = {row[0]: row[1] for row in cursor}
    # Detections per location
    cursor = conn.execute("SELECT location, COUNT(*) as count FROM logs WHERE name != 'Unknown' GROUP BY location")
    location_counts = {row[0]: row[1] for row in cursor}
    # Recent activity
    cursor = conn.execute("SELECT name, location, time FROM logs WHERE name != 'Unknown' ORDER BY time DESC LIMIT 5")
    recent_logs = [{"name": row[0], "location": row[1], "time": row[2]} for row in cursor]
    conn.close()
    
    return render_template("admin_dashboard.html", 
                         floor_data=location_data, 
                         attendance=attendance,
                         person_counts=person_counts, 
                         floor_counts=location_counts, 
                         recent_logs=recent_logs,
                         people=people)

@app.route("/admin/approvals")
@login_required
def admin_approvals():
    if current_user.username != "admin":
        return redirect(url_for("employee_dashboard"))
    return render_template("admin_approvals.html")

@app.route("/admin/employees")
@login_required
def admin_employees():
    if current_user.username != "admin":
        return redirect(url_for("employee_dashboard"))
    return render_template("admin_employees.html")

@app.route("/employee")
@login_required
def employee_dashboard():
    # Normal employee portal (non-HSE)
    return render_template("employee_dashboard.html")

@app.route("/employee/timesheet")
@login_required
def employee_timesheet():
    return render_template("employee_timesheet.html")

@app.route("/employee/requests")
@login_required
def employee_requests():
    return render_template("employee_requests.html")

# ---------------- Admin utility pages ----------------
@app.route("/admin/settings")
@login_required
def admin_settings():
    if current_user.username != "admin":
        return redirect(url_for("admin_hse_settings"))
    return render_template("admin_settings.html")

@app.route("/profile")
@login_required
def profile():
    return render_template("profile.html")

# ---------------- HSE Dashboards ----------------

@app.route("/hse/admin")
@login_required
def admin_hse_dashboard():
    # Only HSE admin user should view this; others go to their HSE employee view
    if current_user.username != "admin_hse":
        return redirect(url_for("employee_hse_dashboard"))
    return render_template("admin_hse_dashboard.html", current_year=datetime.now().year)

@app.route("/hse/admin/reports")
@login_required
def admin_hse_reports():
    if current_user.username != "admin_hse":
        return redirect(url_for("employee_hse_dashboard"))
    return render_template("admin_hse_reports.html")

@app.route("/hse/admin/tracking")
@login_required
def admin_hse_tracking():
    if current_user.username != "admin_hse":
        return redirect(url_for("employee_hse_dashboard"))
    
    # Sample tracking data
    tracking_data = [
        {"date": "2025-01-01", "month": "2025-01", "employee_name": "Employee 19", "department": "Maintenance", "type": "Violation", "description": "Violation related to lifting", "priority": "Medium", "category": "Lifting", "corrective_action": "Site inspection", "action_status": "Closed", "closure_date": "2025-01-12", "days_to_close": 11.0, "reported_by": "Supervisor", "verified_by": "HSE Lead"},
        {"date": "2025-01-02", "month": "2025-01", "employee_name": "Employee 08", "department": "Operations", "type": "Near Miss", "description": "Near Miss related to housekeeping", "priority": "Low", "category": "Housekeeping", "corrective_action": "Toolbox talk", "action_status": "Closed", "closure_date": "2025-01-05", "days_to_close": 3.0, "reported_by": "Supervisor", "verified_by": "HSE Lead"},
        {"date": "2025-01-03", "month": "2025-01", "employee_name": "Employee 15", "department": "Operations", "type": "Near Miss", "description": "Near Miss related to housekeeping", "priority": "High", "category": "Housekeeping", "corrective_action": "JSA review", "action_status": "Closed", "closure_date": "2025-01-21", "days_to_close": 18.0, "reported_by": "Supervisor", "verified_by": "HSE Lead"},
        {"date": "2025-01-04", "month": "2025-01", "employee_name": "Employee 07", "department": "Drilling", "type": "Observation", "description": "Observation related to confined space", "priority": "Low", "category": "Confined Space", "corrective_action": "Corrective training", "action_status": "Open", "closure_date": None, "days_to_close": None, "reported_by": "Supervisor", "verified_by": "HSE Lead"},
        {"date": "2025-01-05", "month": "2025-01", "employee_name": "Employee 12", "department": "HSE", "type": "Violation", "description": "Violation related to behavior", "priority": "Medium", "category": "Behavior", "corrective_action": "PPE refresher", "action_status": "In Progress", "closure_date": None, "days_to_close": None, "reported_by": "Supervisor", "verified_by": "HSE Lead"},
    ]
    # Add more sample data to reach ~60-70 records
    for i in range(6, 70):
        emp_num = (i % 24) + 1
        depts = ["Drilling", "Operations", "Production", "HSE", "Maintenance", "Logistics"]
        types = ["Violation", "Near Miss", "Observation"]
        priorities = ["High", "Medium", "Low"]
        categories = ["Lifting", "Housekeeping", "Behavior", "Environmental", "Electrical", "Confined Space", "PPE", "Process Safety"]
        actions = ["Site inspection", "Toolbox talk", "PPE refresher", "JSA review", "Corrective training"]
        statuses = ["Closed", "In Progress", "Open"]
        
        date_obj = datetime(2025, 1, 1) + timedelta(days=i-1)
        date_str = date_obj.strftime("%Y-%m-%d")
        month_str = date_obj.strftime("%Y-%m")
        
        dept = depts[i % len(depts)]
        type_val = types[i % len(types)]
        priority = priorities[i % len(priorities)]
        category = categories[i % len(categories)]
        action = actions[i % len(actions)]
        status = statuses[i % len(statuses)]
        
        closure_date = None
        days_to_close = None
        if status == "Closed":
            closure_date_obj = date_obj + timedelta(days=5 + (i % 15))
            closure_date = closure_date_obj.strftime("%Y-%m-%d")
            days_to_close = (closure_date_obj - date_obj).days
        
        tracking_data.append({
            "date": date_str,
            "month": month_str,
            "employee_name": f"Employee {emp_num:02d}",
            "department": dept,
            "type": type_val,
            "description": f"{type_val} related to {category.lower()}",
            "priority": priority,
            "category": category,
            "corrective_action": action,
            "action_status": status,
            "closure_date": closure_date,
            "days_to_close": days_to_close,
            "reported_by": "Supervisor",
            "verified_by": "HSE Lead"
        })
    
    return render_template("admin_hse_tracking.html", tracking_data=tracking_data)

@app.route("/hse/admin/tracking/export/excel")
@login_required
def export_hse_tracking_excel():
    if current_user.username != "admin_hse":
        return redirect(url_for("employee_hse_dashboard"))
    
    # Get the same tracking data (same logic as in route)
    tracking_data = []
    for i in range(1, 70):
        emp_num = (i % 24) + 1
        depts = ["Drilling", "Operations", "Production", "HSE", "Maintenance", "Logistics"]
        types = ["Violation", "Near Miss", "Observation"]
        priorities = ["High", "Medium", "Low"]
        categories = ["Lifting", "Housekeeping", "Behavior", "Environmental", "Electrical", "Confined Space", "PPE", "Process Safety"]
        actions = ["Site inspection", "Toolbox talk", "PPE refresher", "JSA review", "Corrective training"]
        statuses = ["Closed", "In Progress", "Open"]
        
        date_obj = datetime(2025, 1, 1) + timedelta(days=i-1)
        date_str = date_obj.strftime("%Y-%m-%d")
        month_str = date_obj.strftime("%Y-%m")
        
        dept = depts[i % len(depts)]
        type_val = types[i % len(types)]
        priority = priorities[i % len(priorities)]
        category = categories[i % len(categories)]
        action = actions[i % len(actions)]
        status = statuses[i % len(statuses)]
        
        closure_date = None
        days_to_close = None
        if status == "Closed":
            closure_date_obj = date_obj + timedelta(days=5 + (i % 15))
            closure_date = closure_date_obj.strftime("%Y-%m-%d")
            days_to_close = (closure_date_obj - date_obj).days
        
        tracking_data.append({
            "date": date_str,
            "month": month_str,
            "employee_name": f"Employee {emp_num:02d}",
            "department": dept,
            "type": type_val,
            "description": f"{type_val} related to {category.lower()}",
            "priority": priority,
            "category": category,
            "corrective_action": action,
            "action_status": status,
            "closure_date": closure_date or "",
            "days_to_close": days_to_close or "",
            "reported_by": "Supervisor",
            "verified_by": "HSE Lead"
        })
    
    # Use openpyxl if available for colored Excel export
    if OPENPYXL_AVAILABLE:
        wb = Workbook()
        ws = wb.active
        ws.title = "HSE Tracking"
        
        # Define colors
        priority_colors = {
            "High": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # Red
            "Medium": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),  # Orange
            "Low": PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
        }
        
        # Headers
        headers = ["Date", "Month", "Employee Name", "Function/Department", "Type", "Description", "Priority", "Category", "Corrective Action Taken", "Action Status", "Closure Date", "Days to Close", "Reported By", "Verified By"]
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data with colors
        for record in tracking_data:
            row = [
                record["date"],
                record["month"],
                record["employee_name"],
                record["department"],
                record["type"],
                record["description"],
                record["priority"],
                record["category"],
                record["corrective_action"],
                record["action_status"],
                record["closure_date"],
                record["days_to_close"],
                record["reported_by"],
                record["verified_by"]
            ]
            ws.append(row)
            
            # Color Priority column (column G, index 6)
            priority_cell = ws.cell(row=ws.max_row, column=7)
            priority_value = record["priority"]
            if priority_value in priority_colors:
                priority_cell.fill = priority_colors[priority_value]
                priority_cell.font = Font(bold=True, color="FFFFFF")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment;filename=HSE_Tracking_Export.xlsx"}
        )
        return response
    else:
        # Fallback to CSV if openpyxl not available
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(["Date", "Month", "Employee Name", "Function/Department", "Type", "Description", "Priority", "Category", "Corrective Action Taken", "Action Status", "Closure Date", "Days to Close", "Reported By", "Verified By"])
        
        # Write data
        for record in tracking_data:
            writer.writerow([
                record["date"],
                record["month"],
                record["employee_name"],
                record["department"],
                record["type"],
                record["description"],
                record["priority"],
                record["category"],
                record["corrective_action"],
                record["action_status"],
                record["closure_date"],
                record["days_to_close"],
                record["reported_by"],
                record["verified_by"]
            ])
        
        response = Response(
            output.getvalue(),
            mimetype="text/csv; charset=utf-8-sig",
            headers={"Content-Disposition": "attachment;filename=HSE_Tracking_Export.csv"}
        )
        return response

@app.route("/hse/admin/monthly-dashboard")
@login_required
def admin_hse_monthly_dashboard():
    if current_user.username != "admin_hse":
        return redirect(url_for("employee_hse_dashboard"))
    
    # Get month from query parameter or use current month
    month_str = request.args.get('month', datetime.now().strftime('%Y-%m'))
    year, month = [int(x) for x in month_str.split('-')]
    
    # Calculate KPI Summary from tracking data
    # In production, fetch from database filtered by month
    total_events = 70  # Sample data
    closed_count = 45
    closed_percent = (closed_count / total_events * 100) if total_events > 0 else 0
    avg_days_to_close = 12.5
    open_in_prog = total_events - closed_count
    
    # KPI Summary by Department
    departments = ["Drilling", "Operations", "Production", "HSE", "Maintenance", "Logistics"]
    kpi_by_dept = []
    for dept in departments:
        dept_total = random.randint(5, 20)
        dept_closed = random.randint(3, dept_total)
        dept_percent = (dept_closed / dept_total * 100) if dept_total > 0 else 0
        dept_days = round(random.uniform(8, 18), 1)
        dept_open = dept_total - dept_closed
        kpi_by_dept.append({
            "department": dept,
            "total_events": dept_total,
            "closed": dept_closed,
            "closed_percent": dept_percent,
            "days_to_close": dept_days,
            "open_in_prog": dept_open
        })
    
    # Employee data with violations, near misses, observations
    employee_stats = {}
    for dept in departments:
        employee_stats[dept] = {}
    
    # Sample employee data - in production, calculate from tracking data
    employees_list = [
        {
            "name": f"Employee {i:02d}", 
            "department": departments[(i-1) % len(departments)],
            "violations": random.randint(0, 4),
            "near_misses": random.randint(0, 4),
            "observations": random.randint(0, 3)
        }
        for i in range(1, 25)
    ]
    
    for emp in employees_list:
        dept = emp["department"]
        if dept not in employee_stats:
            employee_stats[dept] = {}
        if emp["name"] not in employee_stats[dept]:
            employee_stats[dept][emp["name"]] = {
                "violations": emp["violations"], 
                "near_misses": emp["near_misses"], 
                "observations": emp["observations"]
            }
    
    # Priority by Department
    priority_by_dept = {
        "Operations": {"high": 0, "medium": 2, "low": 2},
        "Drilling": {"high": 1, "medium": 6, "low": 0},
        "Maintenance": {"high": 0, "medium": 1, "low": 2},
        "HSE": {"high": 2, "medium": 4, "low": 3},
        "Logistics": {"high": 0, "medium": 0, "low": 0},
        "Production": {"high": 1, "medium": 3, "low": 0},
    }
    
    return render_template("admin_hse_monthly_dashboard.html", 
                         month=month_str,
                         total_events=total_events,
                         closed_count=closed_count,
                         closed_percent=closed_percent,
                         avg_days_to_close=avg_days_to_close,
                         open_in_prog=open_in_prog,
                         employees_list=employees_list,
                         employee_stats=employee_stats,
                         priority_by_dept=priority_by_dept,
                         kpi_by_dept=kpi_by_dept)

@app.route("/hse/admin/employee-report")
@login_required
def employee_hse_report():
    if current_user.username != "admin_hse":
        return redirect(url_for("employee_hse_dashboard"))
    
    employee_id = request.args.get('employee_id', 'E001')
    year = request.args.get('year', datetime.now().year)
    
    # Get employee data
    employee = get_employee_by_id(employee_id)
    if not employee:
        flash("Employee not found", "error")
        return redirect(url_for("admin_hse_employees"))
    
    employee_name = employee["name"]
    employee_photo = employee.get("photo_url", "https://i.pravatar.cc/150?img=1")
    
    # Generate monthly trend data for the selected year
    monthly_trend = []
    for month in range(1, 13):
        month_str = f"{year}-{month:02d}"
        monthly_trend.append({
            "month": month_str,
            "violations": random.randint(0, 4),
            "near_misses": random.randint(0, 2),
            "observations": random.randint(0, 2)
        })
    
    return render_template("employee_hse_report.html",
                         employee_id=employee_id,
                         employee_name=employee_name,
                         employee_photo=employee_photo,
                         year=year,
                         monthly_trend=monthly_trend)

@app.route("/hse/admin/employee-report/export/excel")
@login_required
def export_employee_hse_report_excel():
    if current_user.username != "admin_hse":
        return redirect(url_for("employee_hse_dashboard"))
    
    employee_id = request.args.get('employee_id', 'E001')
    year = request.args.get('year', datetime.now().year)
    
    # Get employee data
    employee = get_employee_by_id(employee_id)
    if not employee:
        flash("Employee not found", "error")
        return redirect(url_for("admin_hse_employees"))
    
    employee_name = employee["name"]
    
    # Generate monthly trend data for the selected year
    monthly_trend = []
    for month in range(1, 13):
        month_str = f"{year}-{month:02d}"
        monthly_trend.append({
            "month": month_str,
            "violations": random.randint(0, 4),
            "near_misses": random.randint(0, 2),
            "observations": random.randint(0, 2)
        })
    
    # Create CSV output
    output = StringIO()
    writer = csv.writer(output)
    
    # Write Monthly Trend Table
    writer.writerow(["Monthly Trend (Selected Year)"])
    writer.writerow(["Month", "Violations", "Near Misses", "Observations"])
    for trend in monthly_trend:
        writer.writerow([
            trend["month"],
            trend["violations"],
            trend["near_misses"],
            trend["observations"]
        ])
    
    response = Response(
        output.getvalue(),
        mimetype="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f"attachment;filename=Employee_HSE_Report_{employee_id}_{year}.csv"}
    )
    
    return response

@app.route("/hse/admin/employee-report/export/pdf")
@login_required
def export_employee_hse_report_pdf():
    if current_user.username != "admin_hse":
        return redirect(url_for("employee_hse_dashboard"))
    
    employee_id = request.args.get('employee_id', 'E001')
    year = request.args.get('year', datetime.now().year)
    
    # Get employee data
    employee = get_employee_by_id(employee_id)
    if not employee:
        flash("Employee not found", "error")
        return redirect(url_for("admin_hse_employees"))
    
    employee_name = employee["name"]
    employee_photo_url = employee.get("photo_url", "https://i.pravatar.cc/150?img=1")
    
    # Generate monthly trend data for the selected year
    monthly_trend = []
    for month in range(1, 13):
        month_str = f"{year}-{month:02d}"
        monthly_trend.append({
            "month": month_str,
            "violations": random.randint(0, 4),
            "near_misses": random.randint(0, 2),
            "observations": random.randint(0, 2)
        })
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, 
                            rightMargin=0.75*inch, leftMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()
    
    # Create custom styles for better formatting
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=20,
        textColor=colors.HexColor('#009879'),
        spaceAfter=20,
        alignment=1,  # Center alignment
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#333333'),
        spaceAfter=12,
        spaceBefore=12,
        fontName='Helvetica-Bold'
    )
    
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#333333'),
        leading=16,
        fontName='Helvetica'
    )
    
    elements = []
    
    # Download employee photo
    try:
        photo_response = requests.get(employee_photo_url, timeout=5)
        if photo_response.status_code == 200:
            photo_buffer = BytesIO(photo_response.content)
            # Make photo circular/square with border
            photo_img = Image(photo_buffer, width=1.8*inch, height=1.8*inch, kind='proportional')
        else:
            photo_img = None
    except:
        photo_img = None
    
    # Header with employee info - improved design
    from reportlab.platypus import Table as RTable
    from reportlab.platypus import KeepTogether
    
    # Create a better header layout
    header_elements = []
    
    # Employee photo in a bordered box
    if photo_img:
        photo_cell = [photo_img]
    else:
        photo_cell = [Paragraph("", info_style)]
    
    # Employee info with better formatting - escape HTML special characters
    from html import escape
    escaped_name = escape(str(employee_name))
    info_text = f"<b>Employee Name:</b> {escaped_name}<br/><b>Employee ID:</b> {employee_id}<br/><b>Year:</b> {year}"
    info_cell = [Paragraph(info_text, info_style)]
    
    header_data = [[photo_cell[0], info_cell[0]]]
    
    header_table = RTable(header_data, colWidths=[2.2*inch, 4.3*inch])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (0, 0), 10),
        ('RIGHTPADDING', (0, 0), (0, 0), 10),
        ('LEFTPADDING', (1, 0), (1, 0), 20),
        ('TOPPADDING', (0, 0), (-1, -1), 15),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8f9fa')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 20))
    
    # Title - centered and styled
    title = Paragraph(f"Employee HSE Report ({year})", title_style)
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Monthly Trend Table
    table_data = [["Month", "Violations", "Near Misses", "Observations"]]
    for trend in monthly_trend:
        table_data.append([
            trend["month"],
            str(trend["violations"]),
            str(trend["near_misses"]),
            str(trend["observations"])
        ])
    
    table = Table(table_data, colWidths=[1.5*inch, 1.3*inch, 1.3*inch, 1.3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#009879')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor('#f8f9fa')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(Paragraph("Monthly Trend (Selected Year)", heading_style))
    elements.append(Spacer(1, 8))
    elements.append(table)
    elements.append(Spacer(1, 20))
    
    # Create chart
    fig, ax = plt.subplots(figsize=(8, 4))
    months = [t["month"] for t in monthly_trend]
    violations = [t["violations"] for t in monthly_trend]
    near_misses = [t["near_misses"] for t in monthly_trend]
    observations = [t["observations"] for t in monthly_trend]
    
    ax.plot(months, violations, marker='o', label='Violations', color='#3b82f6', linewidth=2)
    ax.plot(months, near_misses, marker='s', label='Near Misses', color='#10b981', linewidth=2)
    ax.plot(months, observations, marker='^', label='Observations', color='#ef4444', linewidth=2)
    
    ax.set_xlabel('Month', fontsize=10)
    ax.set_ylabel('Count', fontsize=10)
    ax.set_title('Monthly Trend Chart', fontsize=12, fontweight='bold')
    ax.legend()
    ax.grid(True, alpha=0.3)
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    
    # Save chart to BytesIO
    chart_buffer = BytesIO()
    plt.savefig(chart_buffer, format='png', dpi=150, bbox_inches='tight')
    chart_buffer.seek(0)
    plt.close()
    
    # Add chart to PDF
    elements.append(Paragraph("Monthly Trend Chart", heading_style))
    elements.append(Spacer(1, 8))
    chart_img = Image(chart_buffer, width=6.5*inch, height=3.5*inch)
    elements.append(chart_img)
    elements.append(Spacer(1, 20))
    
    # Build PDF
    doc.build(elements)
    
    response = Response(
        buffer.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment;filename=Employee_HSE_Report_{employee_id}_{year}.pdf"}
    )
    
    return response

@app.route("/hse/admin/monthly-dashboard/export/excel")
@login_required
def export_hse_monthly_dashboard_excel():
    if current_user.username != "admin_hse":
        return redirect(url_for("employee_hse_dashboard"))
    
    month_str = request.args.get('month', datetime.now().strftime('%Y-%m'))
    
    # Get the same data as in the dashboard
    total_events = 70
    closed_count = 45
    closed_percent = (closed_count / total_events * 100) if total_events > 0 else 0
    avg_days_to_close = 12.5
    open_in_prog = total_events - closed_count
    
    employees_list = [
        {"name": f"Employee {i:02d}", "department": ["Drilling", "Operations", "Production", "HSE", "Maintenance", "Logistics"][(i-1) % 6]}
        for i in range(1, 25)
    ]
    
    priority_by_dept = {
        "Operations": {"high": 0, "medium": 2, "low": 2},
        "Drilling": {"high": 1, "medium": 6, "low": 0},
        "Maintenance": {"high": 0, "medium": 1, "low": 2},
        "HSE": {"high": 2, "medium": 4, "low": 3},
        "Logistics": {"high": 0, "medium": 0, "low": 0},
        "Production": {"high": 1, "medium": 3, "low": 0},
    }
    
    # Create CSV output
    output = StringIO()
    writer = csv.writer(output)
    
    # Calculate KPI by Department for export
    departments = ["Drilling", "Operations", "Production", "HSE", "Maintenance", "Logistics"]
    kpi_by_dept = []
    for dept in departments:
        dept_total = random.randint(5, 20)
        dept_closed = random.randint(3, dept_total)
        dept_percent = (dept_closed / dept_total * 100) if dept_total > 0 else 0
        dept_days = round(random.uniform(8, 18), 1)
        dept_open = dept_total - dept_closed
        kpi_by_dept.append({
            "department": dept,
            "total_events": dept_total,
            "closed": dept_closed,
            "closed_percent": dept_percent,
            "days_to_close": dept_days,
            "open_in_prog": dept_open
        })
    
    # Write KPI Summary Table (Horizontal format with multiple rows)
    writer.writerow(["KPI Summary"])
    writer.writerow([
        "Total Events (Month)",
        "Closed (Month)",
        "% Closed (Month)",
        "Days Close (Month)",
        "Open+InProg",
        "Closed",
        "% Closed",
        "Avg Days to Close"
    ])
    # Write data for each department
    for kpi in kpi_by_dept:
        writer.writerow([
            kpi["total_events"],
            kpi["closed"],
            f"{kpi['closed_percent']:.1f}%",
            f"{kpi['days_to_close']:.1f}",
            kpi["open_in_prog"],
            kpi["closed"],
            f"{kpi['closed_percent']:.1f}%",
            f"{kpi['days_to_close']:.1f}"
        ])
    # Write total row
    writer.writerow([
        total_events,
        closed_count,
        f"{closed_percent:.1f}%",
        f"{avg_days_to_close:.1f}",
        open_in_prog,
        closed_count,
        f"{closed_percent:.1f}%",
        f"{avg_days_to_close:.1f}"
    ])
    writer.writerow([])
    
    # Write Priority by Department Table
    writer.writerow(["Priority by"])
    writer.writerow(["Observations", "Department", "High", "Medium", "Low"])
    for dept, priorities in priority_by_dept.items():
        writer.writerow([0, dept, priorities["high"], priorities["medium"], priorities["low"]])
    
    response = Response(
        output.getvalue(),
        mimetype="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f"attachment;filename=HSE_Monthly_Dashboard_{month_str}.csv"}
    )
    
    return response

@app.route("/hse/admin/monthly-dashboard/export/pdf")
@login_required
def export_hse_monthly_dashboard_pdf():
    if current_user.username != "admin_hse":
        return redirect(url_for("employee_hse_dashboard"))
    
    month_str = request.args.get('month', datetime.now().strftime('%Y-%m'))
    
    # Get the same data as in the dashboard
    total_events = 70
    closed_count = 45
    closed_percent = (closed_count / total_events * 100) if total_events > 0 else 0
    avg_days_to_close = 12.5
    open_in_prog = total_events - closed_count
    
    departments = ["Drilling", "Operations", "Production", "HSE", "Maintenance", "Logistics"]
    kpi_by_dept = []
    for dept in departments:
        dept_total = random.randint(5, 20)
        dept_closed = random.randint(3, dept_total)
        dept_percent = (dept_closed / dept_total * 100) if dept_total > 0 else 0
        dept_days = round(random.uniform(8, 18), 1)
        dept_open = dept_total - dept_closed
        kpi_by_dept.append({
            "department": dept,
            "total_events": dept_total,
            "closed": dept_closed,
            "closed_percent": dept_percent,
            "days_to_close": dept_days,
            "open_in_prog": dept_open
        })
    
    priority_by_dept = {
        "Operations": {"high": 0, "medium": 2, "low": 2},
        "Drilling": {"high": 1, "medium": 6, "low": 0},
        "Maintenance": {"high": 0, "medium": 1, "low": 2},
        "HSE": {"high": 2, "medium": 4, "low": 3},
        "Logistics": {"high": 0, "medium": 0, "low": 0},
        "Production": {"high": 1, "medium": 3, "low": 0},
    }
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    title = Paragraph(f"HSE Monthly Dashboard - {month_str}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # KPI Summary Table
    kpi_headers = ["Total Events (Month)", "Closed (Month)", "% Closed (Month)", "Days Close (Month)", 
                   "Open+InProg", "Closed", "% Closed", "Avg Days to Close"]
    kpi_data = [kpi_headers]
    for kpi in kpi_by_dept:
        kpi_data.append([
            str(kpi["total_events"]), str(kpi["closed"]), f"{kpi['closed_percent']:.1f}%",
            f"{kpi['days_to_close']:.1f}", str(kpi["open_in_prog"]), str(kpi["closed"]),
            f"{kpi['closed_percent']:.1f}%", f"{kpi['days_to_close']:.1f}"
        ])
    kpi_data.append([
        str(total_events), str(closed_count), f"{closed_percent:.1f}%",
        f"{avg_days_to_close:.1f}", str(open_in_prog), str(closed_count),
        f"{closed_percent:.1f}%", f"{avg_days_to_close:.1f}"
    ])
    
    kpi_table = Table(kpi_data, colWidths=[1*inch]*8)
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.13, 0.15, 0.24)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.Color(0.95, 0.97, 1.0)]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(Paragraph("KPI Summary", styles['Heading2']))
    elements.append(Spacer(1, 6))
    elements.append(kpi_table)
    elements.append(Spacer(1, 12))
    
    # Priority by Department Table
    priority_data = [["Observations", "Department", "High", "Medium", "Low"]]
    for dept, priorities in priority_by_dept.items():
        priority_data.append([str(0), dept, str(priorities["high"]), str(priorities["medium"]), str(priorities["low"])])
    
    priority_table = Table(priority_data, colWidths=[1*inch, 1.5*inch, 0.8*inch, 0.8*inch, 0.8*inch])
    priority_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.13, 0.15, 0.24)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.Color(0.95, 0.97, 1.0)]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(Paragraph("Priority by Department", styles['Heading2']))
    elements.append(Spacer(1, 6))
    elements.append(priority_table)
    elements.append(Spacer(1, 12))
    
    # Create Priority Breakdown Chart
    fig, ax = plt.subplots(figsize=(8, 5))
    dept_names = list(priority_by_dept.keys())
    high_data = [priority_by_dept[d]["high"] for d in dept_names]
    medium_data = [priority_by_dept[d]["medium"] for d in dept_names]
    low_data = [priority_by_dept[d]["low"] for d in dept_names]
    
    x = range(len(dept_names))
    width = 0.6
    ax.barh(x, high_data, width, label='High', color='#3b82f6')
    ax.barh(x, medium_data, width, left=high_data, label='Medium', color='#ef4444')
    ax.barh(x, low_data, width, left=[h+m for h, m in zip(high_data, medium_data)], label='Low', color='#10b981')
    
    ax.set_yticks(x)
    ax.set_yticklabels(dept_names)
    ax.set_xlabel('Count')
    ax.set_title('Priority Breakdown by Department', fontweight='bold')
    ax.legend()
    ax.grid(True, alpha=0.3, axis='x')
    plt.tight_layout()
    
    chart_buffer = BytesIO()
    plt.savefig(chart_buffer, format='png', dpi=150, bbox_inches='tight')
    chart_buffer.seek(0)
    plt.close()
    
    elements.append(Paragraph("Priority Breakdown Chart", styles['Heading2']))
    elements.append(Spacer(1, 6))
    chart_img = Image(chart_buffer, width=6*inch, height=4*inch)
    elements.append(chart_img)
    
    # Build PDF
    doc.build(elements)
    
    response = Response(
        buffer.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment;filename=HSE_Monthly_Dashboard_{month_str}.pdf"}
    )
    
    return response

@app.route("/hse/admin/tracking/export/pdf")
@login_required
def export_hse_tracking_pdf():
    if current_user.username != "admin_hse":
        return redirect(url_for("employee_hse_dashboard"))
    
    # Get tracking data (same logic as in route)
    tracking_data = []
    for i in range(1, 70):
        emp_num = (i % 24) + 1
        depts = ["Drilling", "Operations", "Production", "HSE", "Maintenance", "Logistics"]
        types = ["Violation", "Near Miss", "Observation"]
        priorities = ["High", "Medium", "Low"]
        categories = ["Lifting", "Housekeeping", "Behavior", "Environmental", "Electrical", "Confined Space", "PPE", "Process Safety"]
        actions = ["Site inspection", "Toolbox talk", "PPE refresher", "JSA review", "Corrective training"]
        statuses = ["Closed", "In Progress", "Open"]
        
        date_obj = datetime(2025, 1, 1) + timedelta(days=i-1)
        date_str = date_obj.strftime("%Y-%m-%d")
        month_str = date_obj.strftime("%Y-%m")
        
        dept = depts[i % len(depts)]
        type_val = types[i % len(types)]
        priority = priorities[i % len(priorities)]
        category = categories[i % len(categories)]
        action = actions[i % len(actions)]
        status = statuses[i % len(statuses)]
        
        closure_date = None
        days_to_close = None
        if status == "Closed":
            closure_date_obj = date_obj + timedelta(days=5 + (i % 15))
            closure_date = closure_date_obj.strftime("%Y-%m-%d")
            days_to_close = (closure_date_obj - date_obj).days
        
        tracking_data.append({
            "date": date_str,
            "month": month_str,
            "employee_name": f"Employee {emp_num:02d}",
            "department": dept,
            "type": type_val,
            "description": f"{type_val} related to {category.lower()}",
            "priority": priority,
            "category": category,
            "corrective_action": action,
            "action_status": status,
            "closure_date": closure_date,
            "days_to_close": days_to_close,
            "reported_by": "Supervisor",
            "verified_by": "HSE Lead"
        })
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, landscape=True)
    styles = getSampleStyleSheet()
    elements = []
    
    title = Paragraph("HSE Tracking Report", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Define colors for Priority and Status
    priority_colors = {
        "High": colors.HexColor("#FF0000"),      # Red
        "Medium": colors.HexColor("#FFA500"),     # Orange
        "Low": colors.HexColor("#00FF00")         # Green
    }
    
    status_colors = {
        "Closed": colors.HexColor("#00FF00"),      # Green
        "In Progress": colors.HexColor("#FFA500"),  # Orange
        "Open": colors.HexColor("#FF0000")       # Red
    }
    
    # Table data with all columns
    headers = ["Date", "Month", "Employee", "Department", "Type", "Priority", "Status"]
    table_data = [headers]
    
    # Limit to 50 rows for PDF
    for record in tracking_data[:50]:
        table_data.append([
            record["date"], 
            record["month"], 
            record["employee_name"], 
            record["department"],
            record["type"], 
            record["priority"], 
            record["action_status"]
        ])
    
    # Create table with appropriate column widths
    col_widths = [0.9*inch, 0.8*inch, 1.0*inch, 0.9*inch, 0.8*inch, 0.8*inch, 0.9*inch]
    table = Table(table_data, colWidths=col_widths)
    
    # Base table style
    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#366092")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]
    
    # Add row backgrounds (alternating)
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            table_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor("#f8f9fa")))
        else:
            table_style.append(('BACKGROUND', (0, i), (-1, i), colors.whitesmoke))
    
    # Add colors for Priority column (column index 5) and Status column (column index 6)
    for i in range(1, len(table_data)):
        priority = table_data[i][5]  # Priority column
        status = table_data[i][6]     # Status column
        
        # Color Priority cell
        if priority in priority_colors:
            table_style.append(('BACKGROUND', (5, i), (5, i), priority_colors[priority]))
            table_style.append(('TEXTCOLOR', (5, i), (5, i), colors.white))
            table_style.append(('FONTNAME', (5, i), (5, i), 'Helvetica-Bold'))
        
        # Color Status cell
        if status in status_colors:
            table_style.append(('BACKGROUND', (6, i), (6, i), status_colors[status]))
            table_style.append(('TEXTCOLOR', (6, i), (6, i), colors.white))
            table_style.append(('FONTNAME', (6, i), (6, i), 'Helvetica-Bold'))
    
    table.setStyle(TableStyle(table_style))
    elements.append(table)
    
    doc.build(elements)
    
    response = Response(
        buffer.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": "attachment;filename=HSE_Tracking_Report.pdf"}
    )
    
    return response

@app.route("/hse/admin/company-dashboard")
@login_required
def hse_company_dashboard():
    if current_user.username != "admin_hse":
        return redirect(url_for("employee_hse_dashboard"))
    
    year = request.args.get('year', datetime.now().year)
    
    # Generate monthly trend data for YTD
    monthly_trend = []
    total_violations = 0
    total_near_misses = 0
    total_observations = 0
    
    for month in range(1, 13):
        violations = random.randint(10, 25)
        near_misses = random.randint(4, 17)
        observations = random.randint(1, 10)
        
        monthly_trend.append({
            "month": month,
            "violations": violations,
            "near_misses": near_misses,
            "observations": observations
        })
        
        total_violations += violations
        total_near_misses += near_misses
        total_observations += observations
    
    # Calculate summary metrics
    total_events = total_violations + total_near_misses + total_observations
    high_percent = 25.9  # Sample data
    closed_percent = 41.0  # Sample data
    avg_days_to_close = 0.0  # Sample data
    
    return render_template("hse_company_dashboard.html",
                         year=year,
                         monthly_trend=monthly_trend,
                         total_violations=total_violations,
                         total_near_misses=total_near_misses,
                         total_observations=total_observations,
                         high_percent=high_percent,
                         closed_percent=closed_percent,
                         avg_days_to_close=avg_days_to_close)

@app.route("/hse/admin/company-dashboard/export/excel")
@login_required
def export_hse_company_dashboard_excel():
    if current_user.username != "admin_hse":
        return redirect(url_for("employee_hse_dashboard"))
    
    year = request.args.get('year', datetime.now().year)
    
    # Generate monthly trend data
    monthly_trend = []
    for month in range(1, 13):
        monthly_trend.append({
            "month": month,
            "violations": random.randint(10, 25),
            "near_misses": random.randint(4, 17),
            "observations": random.randint(1, 10)
        })
    
    # Create CSV output
    output = StringIO()
    writer = csv.writer(output)
    
    # Write Summary Metrics
    writer.writerow(["HSE Company Dashboard (YTD)", year])
    writer.writerow([])
    writer.writerow(["Summary Metrics"])
    writer.writerow(["Total Violations", sum(t["violations"] for t in monthly_trend)])
    writer.writerow(["Total Near Misses", sum(t["near_misses"] for t in monthly_trend)])
    writer.writerow(["Total Observation", sum(t["observations"] for t in monthly_trend)])
    writer.writerow(["High %", "25.9%"])
    writer.writerow(["Closed %", "41.0%"])
    writer.writerow(["Avg Days to Close", "0.0"])
    writer.writerow([])
    
    # Write Monthly Trend Table
    writer.writerow(["Monthly Trend (Events by Type)"])
    writer.writerow(["Month", "Total Violations", "Total Near Misses", "Total Observation"])
    for trend in monthly_trend:
        writer.writerow([
            trend["month"],
            trend["violations"],
            trend["near_misses"],
            trend["observations"]
        ])
    
    response = Response(
        output.getvalue(),
        mimetype="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f"attachment;filename=HSE_Company_Dashboard_{year}.csv"}
    )
    
    return response

@app.route("/hse/admin/company-dashboard/export/pdf")
@login_required
def export_hse_company_dashboard_pdf():
    if current_user.username != "admin_hse":
        return redirect(url_for("employee_hse_dashboard"))
    
    year = request.args.get('year', datetime.now().year)
    
    # Generate monthly trend data
    monthly_trend = []
    for month in range(1, 13):
        monthly_trend.append({
            "month": month,
            "violations": random.randint(10, 25),
            "near_misses": random.randint(4, 17),
            "observations": random.randint(1, 10)
        })
    
    total_violations = sum(t["violations"] for t in monthly_trend)
    total_near_misses = sum(t["near_misses"] for t in monthly_trend)
    total_observations = sum(t["observations"] for t in monthly_trend)
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    title = Paragraph(f"HSE Company Dashboard (YTD) - {year}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Summary Metrics
    summary_data = [
        ["Total Violations", str(total_violations)],
        ["Total Near Misses", str(total_near_misses)],
        ["Total Observation", str(total_observations)],
        ["High %", "25.9%"],
        ["Closed %", "41.0%"],
        ["Avg Days to Close", "0.0"]
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 1.5*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.13, 0.15, 0.24)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.Color(0.95, 0.97, 1.0)]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(Paragraph("Summary Metrics", styles['Heading2']))
    elements.append(Spacer(1, 6))
    elements.append(summary_table)
    elements.append(Spacer(1, 12))
    
    # Monthly Trend Table
    trend_data = [["Month", "Total Violations", "Total Near Misses", "Total Observation"]]
    for trend in monthly_trend:
        trend_data.append([
            str(trend["month"]),
            str(trend["violations"]),
            str(trend["near_misses"]),
            str(trend["observations"])
        ])
    
    trend_table = Table(trend_data, colWidths=[1*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    trend_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.13, 0.15, 0.24)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.Color(0.95, 0.97, 1.0)]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(Paragraph("Monthly Trend (Events by Type)", styles['Heading2']))
    elements.append(Spacer(1, 6))
    elements.append(trend_table)
    elements.append(Spacer(1, 12))
    
    # Create YTD Chart
    fig, ax = plt.subplots(figsize=(8, 5))
    months = [t["month"] for t in monthly_trend]
    violations = [t["violations"] for t in monthly_trend]
    near_misses = [t["near_misses"] for t in monthly_trend]
    observations = [t["observations"] for t in monthly_trend]
    
    ax.plot(months, violations, marker='o', label='Total Violations', color='#3b82f6', linewidth=2)
    ax.plot(months, near_misses, marker='s', label='Total Near Misses', color='#ef4444', linewidth=2)
    ax.plot(months, observations, marker='^', label='Total Observation', color='#10b981', linewidth=2)
    
    ax.set_xlabel('Month', fontsize=10)
    ax.set_ylabel('Count', fontsize=10)
    ax.set_title('YTD', fontsize=12, fontweight='bold')
    ax.legend()
    ax.grid(True, alpha=0.3)
    ax.set_xticks(months)
    plt.tight_layout()
    
    chart_buffer = BytesIO()
    plt.savefig(chart_buffer, format='png', dpi=150, bbox_inches='tight')
    chart_buffer.seek(0)
    plt.close()
    
    elements.append(Paragraph("YTD Chart", styles['Heading2']))
    elements.append(Spacer(1, 6))
    chart_img = Image(chart_buffer, width=6*inch, height=4*inch)
    elements.append(chart_img)
    
    # Build PDF
    doc.build(elements)
    
    response = Response(
        buffer.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment;filename=HSE_Company_Dashboard_{year}.pdf"}
    )
    
    return response

@app.route("/hse/admin/employees")
@login_required
def admin_hse_employees():
    if current_user.username != "admin_hse":
        return redirect(url_for("employee_hse_dashboard"))
    
    # Reset cache to ensure fresh data with English names only
    reset_employee_data_cache()
    
    # Get employee data with real names and photos
    employees = get_employee_data()
    
    return render_template("admin_hse_employees.html", employees=employees)

@app.route("/hse/admin/employees/export/excel")
@login_required
def export_hse_employees_excel():
    if current_user.username != "admin_hse":
        return redirect(url_for("employee_hse_dashboard"))
    
    # Get employee data with real names
    employees = get_employee_data()
    
    # Create CSV output
    output = StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(["Employee ID", "Employee Name", "Department", "Position", "Supervisor", "Hire Date"])
    
    # Write data
    for emp in employees:
        writer.writerow([emp["id"], emp["name"], emp["department"], emp["position"], emp["supervisor"], emp["hire_date"]])
    
    # Create response with Excel-compatible CSV
    response = Response(
        output.getvalue(),
        mimetype="text/csv; charset=utf-8-sig",  # UTF-8 BOM for Excel compatibility
        headers={"Content-Disposition": "attachment;filename=HSE_Employees_Export.csv"}
    )
    
    return response

@app.route("/hse/admin/settings")
@login_required
def admin_hse_settings():
    if current_user.username != "admin_hse":
        return redirect(url_for("employee_hse_dashboard"))
    return render_template("admin_hse_settings.html")

@app.route("/hse/admin/incidents")
@login_required
def admin_hse_incidents():
    if current_user.username != "admin_hse":
        return redirect(url_for("employee_hse_dashboard"))
    return render_template("admin_hse_incidents.html", current_year=datetime.now().year)

@app.route("/hse/admin/inspections")
@login_required
def admin_hse_inspections():
    if current_user.username != "admin_hse":
        return redirect(url_for("employee_hse_dashboard"))
    return render_template("admin_hse_inspections.html", current_year=datetime.now().year)

@app.route("/hse/admin/risks")
@login_required
def admin_hse_risks():
    if current_user.username != "admin_hse":
        return redirect(url_for("employee_hse_dashboard"))
    return render_template("admin_hse_risks.html", current_year=datetime.now().year)

@app.route("/hse/admin/trainings")
@login_required
def admin_hse_trainings():
    if current_user.username != "admin_hse":
        return redirect(url_for("employee_hse_dashboard"))
    return render_template("admin_hse_trainings.html", current_year=datetime.now().year)

@app.route("/hse/admin/ppe")
@login_required
def admin_hse_ppe():
    if current_user.username != "admin_hse":
        return redirect(url_for("employee_hse_dashboard"))
    return render_template("admin_hse_ppe.html", current_year=datetime.now().year)

@app.route("/hse/admin/environmental")
@login_required
def admin_hse_environmental():
    if current_user.username != "admin_hse":
        return redirect(url_for("employee_hse_dashboard"))
    return render_template("admin_hse_environmental.html", current_year=datetime.now().year)

@app.route("/hse/admin/medical")
@login_required
def admin_hse_medical():
    if current_user.username != "admin_hse":
        return redirect(url_for("employee_hse_dashboard"))
    return render_template("admin_hse_medical.html", current_year=datetime.now().year)

@app.route("/hse/employee")
@login_required
def employee_hse_dashboard():
    return render_template("employee_hse_dashboard.html", current_year=datetime.now().year)

# Quick demo login routes
@app.route("/login/admin")
def quick_login_admin():
    conn = sqlite3.connect("users.db")
    cursor = conn.execute("SELECT id, username FROM users WHERE username = ?", ("admin",))
    row = cursor.fetchone()
    conn.close()
    if row:
        login_user(User(row[0], row[1]))
        return redirect(url_for("hr_home"))
    return redirect(url_for("login"))

@app.route("/login/employee")
def quick_login_employee():
    conn = sqlite3.connect("users.db")
    cursor = conn.execute("SELECT id, username FROM users WHERE username = ?", ("employee",))
    row = cursor.fetchone()
    conn.close()
    if row:
        login_user(User(row[0], row[1]))
        return redirect(url_for("employee_dashboard"))
    return redirect(url_for("login"))

@app.route("/login/admin_hse")
def quick_login_admin_hse():
    conn = sqlite3.connect("users.db")
    cursor = conn.execute("SELECT id, username FROM users WHERE username = ?", ("admin_hse",))
    row = cursor.fetchone()
    conn.close()
    if row:
        login_user(User(row[0], row[1]))
        return redirect(url_for("admin_hse_dashboard"))
    return redirect(url_for("login"))

@app.route("/login/user_hse")
def quick_login_user_hse():
    conn = sqlite3.connect("users.db")
    cursor = conn.execute("SELECT id, username FROM users WHERE username = ?", ("user_hse",))
    row = cursor.fetchone()
    conn.close()
    if row:
        login_user(User(row[0], row[1]))
        return redirect(url_for("employee_hse_dashboard"))
    return redirect(url_for("login"))

@app.route("/employee/self/timesheet")
@login_required
def employee_self_timesheet():
    # Build the same preview page but for the logged-in employee (no admin guard)
    name = current_user.username
    now = datetime.now()
    year, month = now.year, now.month
    last_day = monthrange(year, month)[1]
    year_s, month_s = str(year), f"{month:02d}"
    start_date = f"{year_s}-{month_s}-01"
    end_date = f"{year_s}-{month_s}-{last_day:02d}"
    conn = sqlite3.connect("tracking.db")
    cursor = conn.execute(
        """
        SELECT substr(time,1,10) as date,
               MIN(time) as first_in,
               MAX(time) as last_out
        FROM logs
        WHERE name = ? AND date(time) BETWEEN date(?) AND date(?)
        GROUP BY date
        ORDER BY date
        """,
        (name, start_date, end_date)
    )
    day_map = {row[0]: (row[1], row[2]) for row in cursor.fetchall()}
    conn.close()
    rows = []
    for d in range(1, last_day+1):
        date_str = f"{year_s}-{month_s}-{d:02d}"
        if date_str in day_map:
            first_in, last_out = day_map[date_str]
            rows.append({"date": date_str, "first_in": first_in, "last_out": last_out})
        else:
            rows.append({"date": date_str, "first_in": "-", "last_out": "-"})
    if len(rows) == 0:
        rows = [
            {"date": f"{year_s}-{month_s}-01", "first_in": f"{year_s}-{month_s}-01 08:05:00", "last_out": f"{year_s}-{month_s}-01 16:58:00"}
        ]
    meta = {
        "department": "IT Department",
        "employee_no": 1,
        "designation": "Employee",
        "year": year_s,
        "month": month_s
    }
    return render_template("timesheet_preview.html", name=name, rows=rows, meta=meta)

@app.route("/timesheet/preview/<name>")
@login_required
def timesheet_preview(name):
    # Optional year/month query
    year = request.args.get("year")
    month = request.args.get("month")
    if year and month:
        year = int(year); month = int(month)
    else:
        now = datetime.now(); year, month = now.year, now.month
    last_day = monthrange(year, month)[1]
    year_s, month_s = str(year), f"{month:02d}"
    start_date = f"{year_s}-{month_s}-01"
    end_date = f"{year_s}-{month_s}-{last_day:02d}"
    conn = sqlite3.connect("tracking.db")
    cursor = conn.execute(
        """
        SELECT substr(time,1,10) as date,
               MIN(time) as first_in,
               MAX(time) as last_out
        FROM logs
        WHERE name = ? AND date(time) BETWEEN date(?) AND date(?)
        GROUP BY date
        ORDER BY date
        """,
        (name, start_date, end_date)
    )
    day_map = {row[0]: (row[1], row[2]) for row in cursor.fetchall()}
    conn.close()
    rows = []
    for d in range(1, last_day+1):
        date_str = f"{year_s}-{month_s}-{d:02d}"
        if date_str in day_map:
            first_in, last_out = day_map[date_str]
            rows.append({"date": date_str, "first_in": first_in, "last_out": last_out})
        else:
            rows.append({"date": date_str, "first_in": "-", "last_out": "-"})
    if len(rows) == 0:
        rows = [
            {"date": f"{year_s}-{month_s}-01", "first_in": f"{year_s}-{month_s}-01 08:05:00", "last_out": f"{year_s}-{month_s}-01 16:58:00"}
        ]
    meta = {
        "department": "IT Department",
        "employee_no": 1,
        "designation": "Employee",
        "year": year_s,
        "month": month_s
    }
    return render_template("timesheet_preview.html", name=name, rows=rows, meta=meta)
@app.route("/admin/employee/<name>/timesheet")
@login_required
def admin_employee_timesheet(name):
    if current_user.username != "admin":
        return redirect(url_for("employee_dashboard"))
    # Prepare rows for the current month to preview
    now = datetime.now()
    year, month = now.year, now.month
    last_day = monthrange(year, month)[1]
    year_s, month_s = str(year), f"{month:02d}"
    start_date = f"{year_s}-{month_s}-01"
    end_date = f"{year_s}-{month_s}-{last_day:02d}"
    conn = sqlite3.connect("tracking.db")
    cursor = conn.execute(
        """
        SELECT substr(time,1,10) as date,
               MIN(time) as first_in,
               MAX(time) as last_out
        FROM logs
        WHERE name = ? AND date(time) BETWEEN date(?) AND date(?)
        GROUP BY date
        ORDER BY date
        """,
        (name, start_date, end_date)
    )
    day_map = {row[0]: (row[1], row[2]) for row in cursor.fetchall()}
    conn.close()
    rows = []
    for d in range(1, last_day+1):
        date_str = f"{year_s}-{month_s}-{d:02d}"
        if date_str in day_map:
            first_in, last_out = day_map[date_str]
            rows.append({"date": date_str, "first_in": first_in, "last_out": last_out})
        else:
            rows.append({"date": date_str, "first_in": "-", "last_out": "-"})
    if len(rows) == 0:
        rows = [
            {"date": f"{year_s}-{month_s}-01", "first_in": f"{year_s}-{month_s}-01 08:05:00", "last_out": f"{year_s}-{month_s}-01 16:58:00"}
        ]
    # Simple demo metadata
    meta = {
        "department": "IT Department" if name.lower() in ("dalia", "yousef") else "Operations",
        "employee_no": 8,
        "designation": "Software Developer" if name.lower() in ("dalia", "yousef") else "Employee",
        "year": year_s,
        "month": month_s
    }
    return render_template("admin_employee_timesheet.html", name=name, rows=rows, meta=meta)

@app.route("/report")
@login_required
def emergency_status():
    conn = sqlite3.connect("tracking.db")
    cursor = conn.execute("SELECT DISTINCT name FROM logs WHERE name != 'Unknown'")
    names = [row[0] for row in cursor]
    cursor = conn.execute("SELECT DISTINCT location FROM logs")
    locations = [row[0] for row in cursor]
    cursor = conn.execute("SELECT name, location, MAX(time) as last_seen, image_path FROM logs WHERE name != 'Unknown' GROUP BY name")
    people = [{"name": row[0], "location": row[1], "last_seen": row[2], "image_path": row[3]} for row in cursor]
    conn.close()
    return render_template("status_2.html", people=people, names=names, floors=locations)

@app.route("/delete_report_rows", methods=["POST"])
@login_required
def delete_report_rows():
    names = request.get_json()
    if not names:
        return Response("No rows selected", status=400)

    conn = sqlite3.connect("tracking.db")
    cursor = conn.cursor()
    for name in names:
        cursor.execute("DELETE FROM logs WHERE name = ?", (name,))
    conn.commit()
    conn.close()
    return Response("Rows deleted", status=200)

@app.route("/export_report", methods=["POST"])
@login_required
def export_report():
    # Get format and filtered data
    export_format = request.args.get("format", "csv")
    filtered_data = request.get_json()
    if not filtered_data:
        return Response("No data to export", status=400)

    if export_format == "csv":
        # CSV Export
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=["name", "location", "last_seen", "image_path"])
        writer.writeheader()
        for row in filtered_data:
            writer.writerow({
                "name": row["name"],
                "location": row["location"],
                "last_seen": row["last_seen"],
                "image_path": row["image_path"] if row["image_path"] else "No Image"
            })
        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment;filename=emergency_status_report.csv"}
        )
    else:
        # PDF Export
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        # Title
        styles = getSampleStyleSheet()
        elements.append(Paragraph("Emergency Status Report", styles['Title']))
        elements.append(Spacer(1, 12))

        # Table data
        data = [["Name", "Last Known Location", "Last Seen", "Evidence"]]
        for row in filtered_data:
            image_path = row["image_path"] if row["image_path"] and row["image_path"] != "No Image" else None
            image_cell = "No Image"
            if image_path:
                try:
                    # Construct absolute path
                    abs_path = os.path.join(BASE_DIR, image_path.replace('\\', '/'))
                    if os.path.exists(abs_path):
                        # Open and resize image
                        pil_img = PILImage.open(abs_path)
                        img_width, img_height = pil_img.size
                        # Scale to 50 points wide, maintaining aspect ratio
                        scale = 50 / img_width
                        img_width, img_height = int(img_width * scale), int(img_height * scale)
                        img = Image(abs_path, width=img_width, height=img_height)
                        image_cell = img
                except Exception as e:
                    print(f"Error loading image {image_path}: {e}")
                    image_cell = "No Image"

            data.append([
                row["name"],
                row["location"],
                row["last_seen"],
                image_cell
            ])

        # Create table with adjusted column widths
        table = Table(data, colWidths=[1.5*inch, 1.5*inch, 2*inch, 1*inch])
        header_bg = colors.Color(0.13, 0.15, 0.24)
        row_a = colors.Color(0.97, 0.98, 1.0)
        row_b = colors.Color(0.94, 0.97, 1.0)
        accent = colors.Color(0.13, 0.82, 0.93)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), header_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [row_a, row_b]),
            ('GRID', (0, 0), (-1, -1), 0.6, accent),
        ]))
        elements.append(table)

        # Build PDF
        doc.build(elements)
        return Response(
            buffer.getvalue(),
            mimetype="application/pdf",
            headers={"Content-Disposition": "attachment;filename=emergency_status_report.pdf"}
        )

@app.route("/logs/<name>")
@login_required
def person_logs(name):
    conn = sqlite3.connect("tracking.db")
    cursor = conn.execute("SELECT name, location, time, image_path FROM logs WHERE name = ? ORDER BY time DESC", (name,))
    logs = [{"name": row[0], "location": row[1], "time": row[2], "image_path": row[3]} for row in cursor]
    conn.close()
    return render_template("person_logs_2.html", name=name, logs=logs)

@app.route("/delete_log_rows/<name>", methods=["POST"])
@login_required
def delete_log_rows(name):
    entries = request.get_json()
    if not entries:
        return Response("No rows selected", status=400)

    conn = sqlite3.connect("tracking.db")
    cursor = conn.cursor()
    for entry in entries:
        cursor.execute("DELETE FROM logs WHERE name = ? AND location = ? AND time = ?",
                       (entry["name"], entry["location"], entry["time"]))
    conn.commit()
    conn.close()
    return Response("Rows deleted", status=200)

@app.route("/export_logs/<name>")
@login_required
def export_logs(name):
    export_format = request.args.get("format", "csv")
    conn = sqlite3.connect("tracking.db")
    cursor = conn.execute("SELECT name, location, time, image_path FROM logs WHERE name = ? ORDER BY time DESC", (name,))
    logs = [{"name": row[0], "location": row[1], "time": row[2], "image_path": row[3] if row[3] else "No Image"} for row in cursor]
    conn.close()

    # Fallback demo data so PDF/CSV show structure nicely
    if not logs:
        demo_name = name if name else "Employee"
        logs = [
            {"name": demo_name, "location": "Main Room", "time": "2025-11-03 08:03:00", "image_path": "No Image"},
            {"name": demo_name, "location": "Main Room", "time": "2025-11-03 17:01:00", "image_path": "No Image"},
            {"name": demo_name, "location": "Meeting Room", "time": "2025-11-04 08:19:00", "image_path": "No Image"},
            {"name": demo_name, "location": "Meeting Room", "time": "2025-11-04 16:55:00", "image_path": "No Image"}
        ]

    if export_format == "csv":
        # CSV Export
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=["name", "location", "time", "image_path"])
        writer.writeheader()
        writer.writerows(logs)
        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment;filename={name}_logs.csv"}
        )
    else:
        # PDF Export
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        # Title
        styles = getSampleStyleSheet()
        elements.append(Paragraph(f"Logs for {name}", styles['Title']))
        elements.append(Spacer(1, 12))

        # Table data
        data = [["Name", "Location", "Time", "Evidence"]]
        for log in logs:
            image_path = log["image_path"] if log["image_path"] and log["image_path"] != "No Image" else None
            image_cell = "No Image"
            if image_path:
                try:
                    # Construct absolute path
                    abs_path = os.path.join(BASE_DIR, image_path.replace('\\', '/'))
                    if os.path.exists(abs_path):
                        # Open and resize image
                        pil_img = PILImage.open(abs_path)
                        img_width, img_height = pil_img.size
                        # Scale to 50 points wide, maintaining aspect ratio
                        scale = 50 / img_width
                        img_width, img_height = int(img_width * scale), int(img_height * scale)
                        img = Image(abs_path, width=img_width, height=img_height)
                        image_cell = img
                except Exception as e:
                    print(f"Error loading image {image_path}: {e}")
                    image_cell = "No Image"

            data.append([
                log["name"],
                log["location"],
                log["time"],
                image_cell
            ])

        # Create table with adjusted column widths and site-like theme colors
        table = Table(data, colWidths=[1.6*inch, 1.6*inch, 2.2*inch, 1.2*inch])
        header_bg = colors.Color(0.13, 0.15, 0.24)  # deep navy like site
        row_a = colors.Color(0.97, 0.98, 1.0)
        row_b = colors.Color(0.94, 0.97, 1.0)
        accent = colors.Color(0.13, 0.82, 0.93)    # cyan accent
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), header_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [row_a, row_b]),
            ('GRID', (0, 0), (-1, -1), 0.6, accent),
            ('BOX', (0,0), (-1,-1), 1, accent)
        ]))
        elements.append(table)

        # Build PDF
        doc.build(elements)
        return Response(
            buffer.getvalue(),
            mimetype="application/pdf",
            headers={"Content-Disposition": f"attachment;filename={name}_logs.pdf"}
        )

@app.route("/export_company_month")
@login_required
def export_company_month():
    year = request.args.get("year")
    month = request.args.get("month")
    export_format = request.args.get("format", "csv")
    if not year or not month:
        conn = sqlite3.connect("tracking.db")
        cur = conn.execute("SELECT substr(MAX(time),1,4), substr(MAX(time),6,2) FROM logs")
        row = cur.fetchone()
        conn.close()
        year, month = row[0], row[1]

    last_day = monthrange(int(year), int(month))[1]
    start_date = f"{year}-{month}-01"
    end_date = f"{year}-{month}-{last_day:02d}"

    conn = sqlite3.connect("tracking.db")
    cursor = conn.execute(
        """
        SELECT name,
               substr(time,1,10) as date,
               MIN(time) as first_in,
               MAX(time) as last_out
        FROM logs
        WHERE date(time) BETWEEN date(?) AND date(?)
          AND name != 'Unknown'
        GROUP BY name, date
        ORDER BY name, date
        """,
        (start_date, end_date)
    )
    rows = cursor.fetchall()
    conn.close()

    # Fallback demo data if DB has no rows for selected month
    if not rows:
        rows = [
            ("Dalia", "2025-11-01", "2025-11-01 08:01:00", "2025-11-01 17:02:00"),
            ("Yousef", "2025-11-01", "2025-11-01 08:03:00", "2025-11-01 16:58:00"),
            ("Mahmoud", "2025-11-02", "2025-11-02 08:19:00", "2025-11-02 16:55:00")
        ]

    if export_format == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(["name", "date", "first_in", "last_out"])
        for r in rows:
            writer.writerow(r)
        filename = f"company_timesheet_{year}-{month}.csv"
        return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": f"attachment;filename={filename}"})
    else:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []
        elements.append(Paragraph(f"Company Timesheet {year}-{month}", styles['Title']))
        elements.append(Spacer(1, 12))
        data = [["Name","Date","First In","Last Out"]]
        for r in rows:
            data.append([r[0], r[1], r[2], r[3]])
        table = Table(data, colWidths=[1.6*inch, 1.2*inch, 2.2*inch, 2.2*inch])
        header_bg = colors.Color(0.13, 0.15, 0.24)
        row_a = colors.Color(0.97, 0.98, 1.0)
        row_b = colors.Color(0.94, 0.97, 1.0)
        accent = colors.Color(0.13, 0.82, 0.93)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), header_bg),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 12),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [row_a, row_b]),
            ('GRID', (0,0), (-1,-1), 0.6, accent),
            ('BOX', (0,0), (-1,-1), 1, accent)
        ]))
        elements.append(table)
        doc.build(elements)
        filename = f"company_timesheet_{year}-{month}.pdf"
        return Response(buffer.getvalue(), mimetype="application/pdf", headers={"Content-Disposition": f"attachment;filename={filename}"})

@app.route("/export_timesheet/<name>")
@login_required
def export_timesheet(name):
    """Export a single employee monthly timesheet (CSV or PDF)."""
    export_format = request.args.get("format", "pdf")
    year = request.args.get("year")
    month = request.args.get("month")
    if not year or not month:
        conn = sqlite3.connect("tracking.db")
        cur = conn.execute("SELECT substr(MAX(time),1,4), substr(MAX(time),6,2) FROM logs WHERE name = ?", (name,))
        row = cur.fetchone()
        conn.close()
        if row and row[0] and row[1]:
            year, month = row[0], row[1]
        else:
            now = datetime.now()
            year, month = str(now.year), f"{now.month:02d}"

    last_day = monthrange(int(year), int(month))[1]
    start_date = f"{year}-{month}-01"
    end_date = f"{year}-{month}-{last_day:02d}"

    conn = sqlite3.connect("tracking.db")
    cursor = conn.execute(
        """
        SELECT substr(time,1,10) as date,
               MIN(time) as first_in,
               MAX(time) as last_out
        FROM logs
        WHERE name = ?
          AND date(time) BETWEEN date(?) AND date(?)
        GROUP BY date
        ORDER BY date
        """,
        (name, start_date, end_date)
    )
    day_map = {row[0]: (row[1], row[2]) for row in cursor.fetchall()}
    conn.close()

    # Build rows for each day in month; mark absent if not present
    rows = []
    present_days = 0
    for d in range(1, last_day + 1):
        date_str = f"{year}-{month}-{d:02d}"
        if date_str in day_map:
            first_in, last_out = day_map[date_str]
            rows.append([date_str, first_in, last_out])
            present_days += 1
        else:
            rows.append([date_str, "-", "-"])

    # Fallback demo if absolutely empty
    if present_days == 0:
        rows = [
            [f"{year}-{month}-01", f"{year}-{month}-01 08:05:00", f"{year}-{month}-01 16:58:00"],
            [f"{year}-{month}-02", f"{year}-{month}-02 08:18:00", f"{year}-{month}-02 16:55:00"],
        ] + [[f"{year}-{month}-{d:02d}", "-", "-"] for d in range(3, min(8, last_day+1))]

    if export_format == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(["date", "first_in", "last_out"]) 
        writer.writerows(rows)
        filename = f"{name}_timesheet_{year}-{month}.csv"
        return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": f"attachment;filename={filename}"})
    else:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []
        elements.append(Paragraph(f"Time Sheet Report - {name} - {month}/{year}", styles['Title']))
        elements.append(Spacer(1, 10))
        legend = Paragraph("<b>Legend:</b> '-' Absent | First In / Last Out times are local", styles['Normal'])
        elements.append(legend)
        elements.append(Spacer(1, 8))
        data = [["Date", "First In", "Last Out"]] + rows
        table = Table(data, colWidths=[1.4*inch, 2.1*inch, 2.1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.Color(0.1, 0.5, 0.8)),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.whitesmoke, colors.Color(0.95,0.97,1.0)]),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('BOX', (0,0), (-1,-1), 1, colors.black),
        ]))
        elements.append(table)
        doc.build(elements)
        filename = f"{name}_timesheet_{year}-{month}.pdf"
        return Response(buffer.getvalue(), mimetype="application/pdf", headers={"Content-Disposition": f"attachment;filename={filename}"})

@app.route("/map")
@login_required
def map():
    conn = sqlite3.connect("tracking.db")
    # Get the latest location for each person
    cursor = conn.execute("""
        SELECT name, location, MAX(time) as last_seen
        FROM logs
        WHERE name != 'Unknown'
        GROUP BY name
    """)
    people = [{"name": row[0], "location": row[1], "last_seen": row[2]} for row in cursor]
    conn.close()
    return render_template("map.html", people=people)

@app.route("/live_stream")
@login_required
def live_stream():
    return render_template("live_stream.html")

# Dictionary to store video capture objects and locks for each camera
video_streams = {}
locks = {}

# Function to initialize video capture for each camera
def init_video_stream(camera_id):
    if camera_id not in video_streams:
        video_streams[camera_id] = cv2.VideoCapture(camera_id - 1)  # Map camera_id to source (0, 1, 2)
        locks[camera_id] = threading.Lock()

@app.route("/video_feed/<int:camera_id>")
@login_required
def video_feed(camera_id):
    def generate(camera_id):
        init_video_stream(camera_id)
        cap = video_streams[camera_id]
        lock = locks[camera_id]
        while True:
            with lock:
                success, frame = cap.read()
                if not success:
                    break
                _, buffer = cv2.imencode('.jpg', frame)
                frame = buffer.tobytes()
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')

    return Response(generate(camera_id), mimetype='multipart/x-mixed-replace; boundary=frame')

# Cleanup function to release video streams on shutdown
def cleanup_video_streams():
    for cap in video_streams.values():
        cap.release()

atexit.register(cleanup_video_streams)

@app.route("/evidence/<path:filename>")
@login_required
def serve_evidence(filename):
    return send_from_directory("evidence", filename)

@app.route("/people_directory")
@login_required
def people_directory():
    conn = sqlite3.connect("tracking.db")
    
    # Get filter parameters
    selected_date = request.args.get("date")
    selected_period = request.args.get("period", "day")  # default to "day"

    # SQL date filtering logic
    date_filter = ""
    params = []

    if selected_date:
        if selected_period == "day":
            date_filter = "AND DATE(time) = ?"
            params = [selected_date]
        elif selected_period == "month":
            date_filter = "AND strftime('%Y-%m', time) = ?"
            params = [selected_date]
        elif selected_period == "quarter":
            if "-Q" in selected_date:
                year, q = selected_date.split("-Q")
                quarter_months = {
                    "1": ["01", "02", "03"],
                    "2": ["04", "05", "06"],
                    "3": ["07", "08", "09"],
                    "4": ["10", "11", "12"]
                }.get(q, [])
                if quarter_months:
                    placeholders = ','.join('?' * len(quarter_months))
                    date_filter = f"AND strftime('%Y', time) = ? AND strftime('%m', time) IN ({placeholders})"
                    params = [year] + quarter_months

    # Get all unique people
    cursor = conn.execute("SELECT DISTINCT name FROM logs WHERE name != 'Unknown'")
    people = []

    # Photo map
    photo_map = {
        "Abdelrahman_image": "abdelrahman.png",
        "Eng.mahmoud": "Eng.mahmoud.png",
        "Mahmoud_Ahmed": "Mahmoud_Ahmed.png",
        "Mostafa": "Mostafa-2.png",
        "mohamed_ragab": "Ragab.png",
        "yousef": "yousef.png",
        "Dalia": "dalia.PNG",
        "Hagar": "hagar.jpeg",
        "Gamila": "Gamila.jpg"
    }

    roles = {
        "Eng.mahmoud": "Office Manager and Drilling Engineer",
        "Dalia": "HR Specialist",
        "Mostafa": "AI Engineer and Head of Software Team",
        "mohamed_ragab": "Software Team",
        "Abdelrahman_image": "Software Team",
        "Mahmoud_Ahmed": "Software Team",
        "Gamila": "Office Girl",
        "yousef": "Finance",
        "Hagar": "Employee"
    }

    for row in cursor:
        name = row[0]
        image = f"formal photos/{photo_map.get(name, 'default.jpg')}"
        role = roles.get(name, "Employee")

        # Calculate filtered hours
        cursor2 = conn.execute(f"""
            SELECT COUNT(DISTINCT SUBSTR(time, 1, 10)) as days_present,
                   AVG(CAST(SUBSTR(time, 12, 2) AS INTEGER)) as avg_hours,
                   COUNT(*) as total_entries
            FROM logs
            WHERE name = ? {date_filter}
        """, (name, *params))
        stats = cursor2.fetchone()
        days_present = stats[0] or 0
        avg_hours = round(stats[1], 1) if stats[1] else 0
        total_entries = stats[2] or 0

        attendance_rate = round((days_present / 20) * 100) if days_present else 0

        # Excuse hours
        if selected_date:
            if selected_period == "day":
                cursor3 = conn.execute(
                    "SELECT SUM(hours) FROM excuses WHERE name = ? AND date = ?",
                    (name, selected_date)
                )
            elif selected_period == "month":
                cursor3 = conn.execute(
                    "SELECT SUM(hours) FROM excuses WHERE name = ? AND strftime('%Y', date) = ?",
                    (name, selected_date)
                )
            elif selected_period == "quarter" and "-Q" in selected_date:
                year, q = selected_date.split("-Q")
                months = {
                    "1": ["01", "02", "03"],
                    "2": ["04", "05", "06"],
                    "3": ["07", "08", "09"],
                    "4": ["10", "11", "12"]
                }.get(q, [])
                placeholders = ','.join('?' * len(months))
                cursor3 = conn.execute(
                    f"SELECT SUM(hours) FROM excuses WHERE name = ? AND strftime('%Y', date) = ? AND strftime('%m', date) IN ({placeholders})",
                    (name, year, *months)
                )
            else:
                cursor3 = conn.execute("SELECT SUM(hours) FROM excuses WHERE name = ?", (name,))
        else:
            cursor3 = conn.execute("SELECT SUM(hours) FROM excuses WHERE name = ?", (name,))

        excuse_hours = cursor3.fetchone()[0] or 0

        # Estimated hours from total entries
        estimated_hours = round(total_entries / 2.0, 1)  # adjust logic as needed

        current_hours = estimated_hours + excuse_hours

        people.append({
            "id": name,
            "name": name,
            "image": image,
            "role": role,
            "attendance_rate": attendance_rate,
            "days_present": days_present,
            "average_hours": avg_hours,
            "current_hours": current_hours
        })

    conn.close()
    return render_template("people_directory.html", people=people, selected_date=selected_date, selected_period=selected_period)

@app.route("/generate_report/<person_id>")
@login_required
def generate_report(person_id):
    conn = sqlite3.connect("tracking.db")

    # Get person's basic info
    cursor = conn.execute("""
        SELECT name, image_path, MAX(time) as last_seen
        FROM logs
        WHERE name = ? AND name != 'Unknown'
        GROUP BY name
    """, (person_id,))
    person = cursor.fetchone()

    if not person:
        conn.close()
        return "Person not found", 404

    name, image_path, last_seen = person

    # Get attendance statistics
    cursor = conn.execute("""
        SELECT COUNT(DISTINCT SUBSTR(time, 1, 10)) as days_present,
               AVG(CAST(SUBSTR(time, 12, 2) AS INTEGER)) as avg_hours,
               MIN(time) as first_seen,
               MAX(time) as last_seen,
               COUNT(*) as total_entries
        FROM logs
        WHERE name = ?
    """, (person_id,))
    stats = cursor.fetchone()
    days_present, avg_hours, first_seen, last_seen, total_entries = stats

    # Get most visited locations
    cursor = conn.execute("""
        SELECT location, COUNT(*) as count
        FROM logs
        WHERE name = ?
        GROUP BY location
        ORDER BY count DESC
        LIMIT 3
    """, (person_id,))
    top_locations = cursor.fetchall()

    # Get daily hours distribution
    cursor = conn.execute("""
        SELECT SUBSTR(time, 1, 10) as date,
               COUNT(*) as entries,
               MAX(CAST(SUBSTR(time, 12, 2) AS INTEGER)) - MIN(CAST(SUBSTR(time, 12, 2) AS INTEGER)) as hours
        FROM logs
        WHERE name = ?
        GROUP BY date
        ORDER BY date DESC
        LIMIT 7
    """, (person_id,))
    daily_hours = cursor.fetchall()

    # Get time distribution
    cursor = conn.execute("""
        SELECT 
            CASE 
                WHEN CAST(SUBSTR(time, 12, 2) AS INTEGER) BETWEEN 6 AND 11 THEN 'Morning'
                WHEN CAST(SUBSTR(time, 12, 2) AS INTEGER) BETWEEN 12 AND 17 THEN 'Afternoon'
                ELSE 'Evening'
            END as time_period,
            COUNT(*) as count
        FROM logs
        WHERE name = ?
        GROUP BY time_period
    """, (person_id,))
    time_distribution = cursor.fetchall()

    # Add role information
    roles = {
        "Eng. Mahmoud": "Office Manager and Drilling Engineer",
        "Dalia": "HR Specialist",
        "Mostafa": "AI Engineer and Head of Software Team",
        "Ragab": "Software Team",
        "Abdelrahman": "Software Team",
        "Mahmoud Ahmed": "Software Team",
        "Gamila": "Office Girl",
        "Yousef": "Finance"
    }
    role = roles.get(name, "Unknown Role")

    conn.close()

    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title = Paragraph(f"Person Report: {name}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))

    # Header Section
    header_data = []
    header_data.append([
        Image(os.path.join('Formal photos', f"{name}.jpg"), width=100, height=100),
        [
            Paragraph(f"<b>Name:</b> {name}", styles['Normal']),
            Paragraph(f"<b>Role:</b> {role}", styles['Normal']),
            Paragraph(f"<b>Issue's Date:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']),
            Paragraph(f"<b>First Seen:</b> {first_seen}", styles['Normal']),
            Paragraph(f"<b>Last Seen:</b> {last_seen}", styles['Normal']),
            Paragraph(f"<b>Total Entries:</b> {total_entries}", styles['Normal'])
        ]
    ])
    header_table = Table(header_data, colWidths=[120, 400])
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('PADDING', (0, 0), (-1, -1), 10)
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 20))

    # Attendance Statistics
    elements.append(Paragraph("Attendance Statistics", styles['Heading2']))
    elements.append(Spacer(1, 10))

    stats_data = [
        ['Days Present', 'Average Hours', 'Total Entries'],
        [str(days_present), f"{avg_hours:.1f}", str(total_entries)]
    ]
    stats_table = Table(stats_data, colWidths=[200, 200, 200])
    stats_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 20))

    # Location Analytics
    elements.append(Paragraph("Location Analytics", styles['Heading2']))
    elements.append(Spacer(1, 10))

    location_data = [['Location', 'Visits']]
    for loc, count in top_locations:
        location_data.append([loc, str(count)])

    location_table = Table(location_data, colWidths=[300, 100])
    location_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(location_table)
    elements.append(Spacer(1, 20))

    # Time Distribution
    elements.append(Paragraph("Time Distribution", styles['Heading2']))
    elements.append(Spacer(1, 10))

    time_data = [['Time Period', 'Entries']]
    for period, count in time_distribution:
        time_data.append([period, str(count)])

    time_table = Table(time_data, colWidths=[200, 200])
    time_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(time_table)
    elements.append(Spacer(1, 20))

    # Recent Activity
    elements.append(Paragraph("Recent Activity (Last 7 Days)", styles['Heading2']))
    elements.append(Spacer(1, 10))

    activity_data = [['Date', 'Entries', 'Hours Present']]
    for date, entries, hours in daily_hours:
        activity_data.append([date, str(entries), f"{hours:.1f}"])

    activity_table = Table(activity_data, colWidths=[150, 150, 150])
    activity_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(activity_table)

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"{name}_report.pdf",
        mimetype='application/pdf'
    )

@app.route("/formal_photos/<path:filename>")
@login_required
def serve_formal_photo(filename):
    return send_from_directory('Formal photos', filename)

@app.route("/excuses", methods=["GET", "POST"])
@login_required
def excuses():
    conn = sqlite3.connect("tracking.db")
    cursor = conn.execute("SELECT DISTINCT name FROM logs WHERE name != 'Unknown'")
    employees = [row[0] for row in cursor]

    if request.method == "POST":
        name = request.form["name"]
        date = request.form["date"]
        hours = request.form["hours"]
        reason = request.form["reason"]
        approved_by = current_user.username
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        conn.execute(
            "INSERT INTO excuses (name, date, hours, reason, approved_by, created_at) VALUES (?, ?, ?, ?, ?, ?)",
            (name, date, hours, reason, approved_by, created_at)
        )
        conn.commit()

    cursor = conn.execute("SELECT rowid as id, name, date, hours, reason, approved_by, created_at FROM excuses")
    excuses = [
        {
            "id": row[0],
            "name": row[1],
            "date": row[2],
            "hours": row[3],
            "reason": row[4],
            "approved_by": row[5],
            "created_at": row[6],
        }
        for row in cursor
    ]
    conn.close()
    return render_template("excuses.html", employees=employees, excuses=excuses)

@app.route("/add_excuse", methods=["POST"])
@login_required
def add_excuse():
    conn = sqlite3.connect("tracking.db")
    name = request.form["name"]
    date = request.form["date"]
    hours = request.form["hours"]
    reason = request.form["reason"]
    approved_by = current_user.username
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn.execute(
        "INSERT INTO excuses (name, date, hours, reason, approved_by, created_at) VALUES (?, ?, ?, ?, ?, ?)",
        (name, date, hours, reason, approved_by, created_at)
    )
    conn.commit()
    conn.close()
    return redirect(url_for("excuses"))

@app.route("/edit_excuse/<int:id>", methods=["GET", "POST"])
@login_required
def edit_excuse(id):
    conn = sqlite3.connect("tracking.db")
    if request.method == "POST":
        name = request.form["name"]
        date = request.form["date"]
        hours = request.form["hours"]
        reason = request.form["reason"]
        conn.execute(
            "UPDATE excuses SET name = ?, date = ?, hours = ?, reason = ? WHERE rowid = ?",
            (name, date, hours, reason, id)
        )
        conn.commit()
        conn.close()
        return redirect(url_for("excuses"))

    cursor = conn.execute("SELECT name, date, hours, reason FROM excuses WHERE rowid = ?", (id,))
    excuse = cursor.fetchone()
    conn.close()
    if not excuse:
        return "Excuse not found", 404

    return render_template("edit_excuse.html", excuse={
        "id": id,
        "name": excuse[0],
        "date": excuse[1],
        "hours": excuse[2],
        "reason": excuse[3]
    })

@app.route("/delete_excuse/<int:id>", methods=["POST"])
@login_required
def delete_excuse(id):
    conn = sqlite3.connect("tracking.db")
    conn.execute("DELETE FROM excuses WHERE rowid = ?", (id,))
    conn.commit()
    conn.close()
    return redirect(url_for("excuses"))

# ---------------- HR Module Routes ----------------
@app.route("/hr")
@app.route("/hr/")
@login_required
def hr_home():
    # Check if user is admin
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("hr/index.html")

@app.route("/hr/employees")
@login_required
def hr_employees():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("hr/employees/index.html")

@app.route("/hr/departments")
@login_required
def hr_departments():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("hr/departments/index.html")

@app.route("/hr/recruitment")
@login_required
def hr_recruitment():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("hr/recruitment/index.html")

@app.route("/hr/attendance")
@login_required
def hr_attendance():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("hr/attendance/index.html")

@app.route("/hr/leaves")
@login_required
def hr_leaves():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("hr/leaves/index.html")

@app.route("/hr/payroll")
@login_required
def hr_payroll():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("hr/payroll/index.html")

@app.route("/hr/compensation")
@login_required
def hr_compensation():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("hr/compensation/index.html")

@app.route("/hr/performance")
@login_required
def hr_performance():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("hr/performance/index.html")

@app.route("/hr/training")
@login_required
def hr_training():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("hr/training/index.html")

@app.route("/hr/sites")
@login_required
def hr_sites():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("hr/sites/index.html")

@app.route("/hr/access")
@login_required
def hr_access():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("hr/access/index.html")

@app.route("/hr/reports")
@login_required
def hr_reports():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("hr/reports/index.html")

# ---------------- Unauthorized Area Dashboard Routes ---------------- 
@app.route("/unauthorized")
@app.route("/unauthorized/")
@login_required
def unauthorized_home():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("unauthorized/index.html")

@app.route("/unauthorized/violations")
@login_required
def unauthorized_violations():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("unauthorized/violations.html")

@app.route("/unauthorized/unauthorized-areas")
@login_required
def unauthorized_areas():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("unauthorized/unauthorized-areas.html")

@app.route("/unauthorized/violations-by-person")
@login_required
def unauthorized_violations_by_person():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("unauthorized/violations-by-person.html")

@app.route("/unauthorized/violations-by-department")
@login_required
def unauthorized_violations_by_department():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("unauthorized/violations-by-department.html")

@app.route("/unauthorized/violations-by-location")
@login_required
def unauthorized_violations_by_location():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("unauthorized/violations-by-location.html")

@app.route("/unauthorized/risk-analysis")
@login_required
def unauthorized_risk_analysis():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("unauthorized/risk-analysis.html")

@app.route("/unauthorized/daily-log")
@login_required
def unauthorized_daily_log():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("unauthorized/daily-log.html")

@app.route("/unauthorized/incident-reporting")
@login_required
def unauthorized_incident_reporting():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("unauthorized/incident-reporting.html")

@app.route("/unauthorized/investigations")
@login_required
def unauthorized_investigations():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("unauthorized/investigations.html")

@app.route("/unauthorized/corrective-actions")
@login_required
def unauthorized_corrective_actions():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("unauthorized/corrective-actions.html")

@app.route("/unauthorized/training-needs")
@login_required
def unauthorized_training_needs():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("unauthorized/training-needs.html")

@app.route("/unauthorized/compliance-score")
@login_required
def unauthorized_compliance_score():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("unauthorized/compliance-score.html")

@app.route("/unauthorized/policies")
@login_required
def unauthorized_policies():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("unauthorized/policies.html")

@app.route("/unauthorized/security-log")
@login_required
def unauthorized_security_log():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("unauthorized/security-log.html")

@app.route("/unauthorized/roles-permissions")
@login_required
def unauthorized_roles_permissions():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("unauthorized/roles-permissions.html")

@app.route("/unauthorized/reports")
@login_required
def unauthorized_reports():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("unauthorized/reports.html")

@app.route("/unauthorized/settings")
@login_required
def unauthorized_settings():
    if current_user.username != "admin":
        flash("Access denied. Admin only.", "error")
        return redirect(url_for("login"))
    return render_template("unauthorized/settings.html")

@app.errorhandler(404)
def page_not_found(e):
    return render_template("404_2.html"), 404

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)